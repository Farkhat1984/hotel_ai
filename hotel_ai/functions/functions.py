import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import uuid


# Создание Excel файла с данными
def create_excel_file(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Первая строка - даты с сегодняшнего дня до конца года
    today = datetime.today()
    end_of_year = datetime(today.year, 12, 31)
    delta = timedelta(days=1)

    dates = []
    while today <= end_of_year:
        dates.append(today.strftime("%d-%m-%Y"))
        today += delta

    ws.append(["Room Type/Date", "Booking ID", "Guest Name"] + dates)

    # Первая колонка - типы номеров и их количество
    room_types = [("Single", 10), ("Double", 10), ("Lux", 5)]
    for room_type, count in room_types:
        for i in range(count):
            ws.append([f"{room_type} {i + 1}"])

    wb.save(file_path)


# Функция для проверки доступности номеров
def check_availability(file_path, room_type, date):
    file_path = "C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx"
    if not os.path.exists(file_path):
        create_excel_file(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    # Найти колонку с нужной датой
    date_col = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == date:
                date_col = cell.column
                break

    if not date_col:
        raise ValueError("Date not found in the sheet.")

    # Проверить доступность номеров
    available_rooms = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        room = row[0].value
        if room and room.startswith(room_type):
            room_cell = ws.cell(row=row[0].row, column=date_col)
            if not room_cell.fill or room_cell.fill.start_color.index == "00000000":
                available_rooms.append(room)

    return available_rooms


# Функция для бронирования номеров
def book_room(file_path, room, date, guest_name):
    file_path = "C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx"
    if not os.path.exists(file_path):
        create_excel_file(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    # Найти колонку с нужной датой
    date_col = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == date:
                date_col = cell.column
                break

    if not date_col:
        raise ValueError("Date not found in the sheet.")

    # Найти нужную комнату
    room_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == room:
            room_row = row[0].row
            break

    if not room_row:
        raise ValueError("Room not found in the sheet.")

    # Забронировать комнату
    room_cell = ws.cell(row=room_row, column=date_col)
    if room_cell.fill and room_cell.fill.start_color.index in ["FFFF00", "00FF00"]:
        raise ValueError("Room is already booked or occupied.")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    room_cell.fill = yellow_fill

    # Добавить уникальный ID бронирования и примечание
    booking_id = str(uuid.uuid4())
    ws.cell(row=room_row, column=2).value = booking_id
    ws.cell(row=room_row, column=3).value = guest_name

    wb.save(file_path)
    return booking_id


# Функция для заселения в номера
def check_in_room(file_path, booking_id, date):
    file_path = "C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx"
    if not os.path.exists(file_path):
        create_excel_file(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    # Найти колонку с нужной датой
    date_col = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == date:
                date_col = cell.column
                break

    if not date_col:
        raise ValueError("Date not found in the sheet.")

    # Найти нужную бронь по ID
    room_row = None
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[1].value == booking_id:
            room_row = row[0].row
            break

    if not room_row:
        raise ValueError("Booking ID not found in the sheet.")

    # Заселить в комнату
    room_cell = ws.cell(row=room_row, column=date_col)
    if room_cell.fill and room_cell.fill.start_color.index == "00FF00":
        raise ValueError("Room is already occupied.")

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    room_cell.fill = green_fill

    wb.save(file_path)


# Функция для выезда до окончания срока
def check_out_room(file_path, booking_id):
    file_path = "C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx"
    if not os.path.exists(file_path):
        create_excel_file(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    # Найти нужную бронь по ID
    room_row = None
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[1].value == booking_id:
            room_row = row[0].row
            break

    if not room_row:
        raise ValueError("Booking ID not found in the sheet.")

    # Освободить комнату
    for cell in ws[room_row]:
        if cell.column > 3:
            cell.fill = PatternFill(start_color="00000000", end_color="00000000", fill_type=None)

    # Удалить ID бронирования и примечание
    ws.cell(row=room_row, column=2).value = None
    ws.cell(row=room_row, column=3).value = None

    wb.save(file_path)


# Функция авто-выселения
def auto_check_out(file_path):
    file_path = "C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx"
    if not os.path.exists(file_path):
        create_excel_file(file_path)
    wb = load_workbook(file_path)
    ws = wb.active

    today = datetime.today().strftime("%Y-%m-%d")

    # Найти колонку с сегодняшней датой
    date_col = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == today:
                date_col = cell.column
                break

    if not date_col:
        raise ValueError("Today's date not found in the sheet.")

    # Авто-выселение
    for row in ws.iter_rows(min_row=2, max_col=2):
        booking_id = row[1].value
        if booking_id:
            room_row = row[0].row
            room_cell = ws.cell(row=room_row, column=date_col)
            if room_cell.fill and room_cell.fill.start_color.index == "00FF00":
                check_out_room(file_path, booking_id)



create_excel_file("C:\\Pycharm\\pythonProject\\hotel_ai\\data\\hotel_bookings.xlsx")